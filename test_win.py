import sys
import time
import pandas as pd
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMainWindow, QAction, QMenu, QApplication, qApp, QFileDialog, QTextEdit, QVBoxLayout, QPushButton, QWidget, QMessageBox


class Example(QMainWindow):

    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        main_frame = QWidget()
        main_frame.setLayout(layout)
        self.setCentralWidget(main_frame)

        menubar = self.menuBar()
        fileMenu_1 = menubar.addMenu("Pre-Processing")
        fileMenu_3 = menubar.addMenu("Exist")

        # file menu 1 setting
        tool_1 = QMenu("设备实时风控", self)
        fileMenu_1.addMenu(tool_1)
        tool_2 = QMenu("车辆实时风控", self)
        fileMenu_1.addMenu(tool_2)

        act_1 = QAction("统计分析", self)
        act_1.triggered.connect(self.trans)
        tool_1.addAction(act_1)
        act_2 = QAction("仅分离", self)
        act_2.triggered.connect(self.trans_simple)
        tool_1.addAction(act_2)
        act_3 = QAction("多表合并", self)
        act_3.triggered.connect(self.combine)
        tool_1.addAction(act_3)
        act_4 = QAction("综合提取", self)
        act_4.triggered.connect(self.picks)
        tool_2.addAction(act_4)
        act_5 = QAction("多表合并", self)
        act_5.triggered.connect(self.combine)
        tool_2.addAction(act_5)

        # file menu 3 setting
        exist_act = QAction("exist", self)
        exist_act.setStatusTip("exist app")
        exist_act.triggered.connect(qApp.quit)
        fileMenu_3.addAction(exist_act)

        self.contents = QTextEdit()
        layout.addWidget(self.contents)
        self.button1 = QPushButton('关闭主窗口')
        self.button1.clicked.connect(self.onButtonClick1)
        self.button2 = QPushButton("清除")
        self.button2.clicked.connect(self.onButtonClick2)
        layout.addWidget(self.button1)
        layout.addWidget(self.button2)

        self.setGeometry(300, 300, 400, 300)
        self.setWindowTitle('Submenu')
        self.show()

    def trans(self):
        # 1.计时器
        # 1.1计时器开始计时
        time1 = time.time()

        # 2.读取设备信息表
        # 2.1选取源文件，源文件即设备信息表
        fileName_choose_origin, filetype = QFileDialog.getOpenFileName(self, "选取源文件", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        self.contents.append(str(fileName_choose_origin))
        if fileName_choose_origin == "" or '设备实时风控' not in fileName_choose_origin:
            print("\n取消选择")
            return
        # 2.2选取字段
        data = pd.read_excel(fileName_choose_origin)
        data = data[['车架号', '设备号', '设备型号', '位置', '在线状态', '报告时间', '行驶状态', '设备离线天数']]
        # 2.3有线无线分离
        data['状态_综合'] = data.apply(lambda x: self.df_func_2(x['在线状态'], x['行驶状态'], x['设备离线天数']), axis=1)
        wire = data[data['设备型号'].str.contains('有线')]
        wireless = data[data['设备型号'].str.contains('无线')]
        # 2.4无线1无线2分离
        wireless_asc = wireless.sort_values(by=['设备号'], ascending=True)
        wireless_desc = wireless.sort_values(by=['设备号'], ascending=False)
        wireless_1 = wireless_asc.drop_duplicates(subset=['车架号'], keep='first', inplace=False)
        wireless_2 = wireless_desc.drop_duplicates(subset=['车架号'], keep='first', inplace=False)
        # 2.5无线汇总
        wireless_cpr = pd.merge(wireless_1, wireless_2, how='left', on=['车架号', '车架号'])
        wireless_cpr = wireless_cpr[['车架号', '设备号_x', '设备号_y', '报告时间_x', '报告时间_y']]
        wireless_cpr['报告时间_综合'] = wireless_cpr.apply(lambda x: self.df_func_1(x['报告时间_x'], x['报告时间_y']), axis=1)
        # 2.6总数计数
        device_count = pd.pivot_table(data, index=['车架号'], values=['设备号'], aggfunc=['count'])
        device_count = device_count.reset_index()
        device_count.columns = device_count.columns.droplevel(1)

        # 3.结果输出
        # 3.1选取目标文件，目标文件即结果追加至的文件
        fileName_choose_target, filetype = QFileDialog.getOpenFileName(self, "选取目标文件", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        self.contents.append(str(fileName_choose_target))
        if fileName_choose_target == "":
            print("\n取消选择")
            return
        book = load_workbook(fileName_choose_target)
        writer = pd.ExcelWriter(fileName_choose_target)
        writer.book = book
        wire.to_excel(writer, "有线设备")
        wireless_1.to_excel(writer, "无线设备1")
        wireless_2.to_excel(writer, "无线设备2")
        wireless_cpr.to_excel(writer, "无线设备综合")
        device_count.to_excel(writer, "设备总数")
        book.close()
        writer.close()

        # 1.2计时结束，显示计时结果同时表示运行成功
        time2 = time.time()
        deltat = round(time2-time1, 4)
        self.show_messagebox(deltat)

    def df_func_1(self, a, b):
        if a >= b:
            return a
        else:
            return b

    def df_func_2(self, a, b, c):
        if a[0:2] == "离线":
            return a[0:2] + c
        elif a == "在线":
            return a[0:2] + b[0:2]
        else:
            return "--"

    def df_func_3(self, a):
        b = a.split('(覆盖半径')[0]
        return b

    def trans_simple(self):
        # 1.计时器
        # 1.1计时器开始计时
        time1 = time.time()

        # 2.读取设备信息表
        # 2.1选取源文件，源文件即设备信息表
        fileName_choose_origin, filetype = QFileDialog.getOpenFileName(self, "选取源文件", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        self.contents.append(str(fileName_choose_origin))
        if fileName_choose_origin == "" or '设备实时风控' not in fileName_choose_origin:
            print("\n取消选择")
            return
        # 2.2全量字段
        data = pd.read_excel(fileName_choose_origin)
        # 2.3有线无线分离
        wire = data[data['设备型号'].str.contains('有线')]
        wireless = data[data['设备型号'].str.contains('无线')]
        # 2.4无线1无线2分离
        wireless_asc = wireless.sort_values(by=['设备号'], ascending=True)
        wireless_desc = wireless.sort_values(by=['设备号'], ascending=False)
        wireless_1 = wireless_asc.drop_duplicates(subset=['车架号'], keep='first', inplace=False)
        wireless_2 = wireless_desc.drop_duplicates(subset=['车架号'], keep='first', inplace=False)

        # 3.结果输出
        # 3.1选取目标文件，目标文件即结果追加至的文件
        fileName_choose_target, filetype = QFileDialog.getOpenFileName(self, "选取目标文件", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        self.contents.append(str(fileName_choose_target))
        if fileName_choose_target == "":
            print("\n取消选择")
            return
        book = load_workbook(fileName_choose_target)
        writer = pd.ExcelWriter(fileName_choose_target)
        writer.book = book
        wire.to_excel(writer, "有线设备")
        wireless_1.to_excel(writer, "无线设备1")
        wireless_2.to_excel(writer, "无线设备2")
        book.close()
        writer.close()

        # 1.2计时结束，显示计时结果同时表示运行成功
        time2 = time.time()
        deltat = round(time2-time1, 4)
        self.show_messagebox(deltat)

    def combine(self):
        # 1.计时器
        # 1.1计时器开始计时
        time1 = time.time()

        # 2.读取数据
        # 2.1选取文件
        fileName_choose_origin, filetype = QFileDialog.getOpenFileNames(self, "选取源文件", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        self.contents.append(str(fileName_choose_origin))
        if fileName_choose_origin == "":
            print("\n取消选择")
            return
        # 2.2合并文件
        df = pd.DataFrame()
        for i in range(1, len(fileName_choose_origin)+1):
            data = pd.read_excel(fileName_choose_origin[i-1])
            df = pd.concat([df, data])

        # 3、结果输出
        fileName_choose_saveas, filetype = QFileDialog.getSaveFileName(self, "保存至", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        df.to_excel(fileName_choose_saveas)

        # 1.2计时结束，显示计时结果同时表示运行成功
        time2 = time.time()
        deltat = round(time2-time1, 4)
        self.show_messagebox(deltat)

    def picks(self):
        fileName_choose_origin, filetype = QFileDialog.getOpenFileName(self, "选取源文件", "C:\\Users\\geste\\Desktop\\", "Text Files (*.xlsx)")
        self.contents.append(str(fileName_choose_origin))
        if fileName_choose_origin == "" or '车辆实时风控' not in fileName_choose_origin:
            print("\n取消选择")
            return
        data = pd.read_excel(fileName_choose_origin)
        data = data[['车架号', '在线状态', '车辆离线天数', '行驶状态', '位置']]
        data['状态综合'] = data.apply(lambda x: self.df_func_2(x['在线状态'], x['行驶状态'], x['车辆离线天数']), axis=1)
        data['位置综合'] = data.apply(lambda x: self.df_func_3(x['位置']), axis=1)

    def show_messagebox(self, message):
        content = "success! it costs " + str(message) + " seconds"
        QMessageBox.information(self, 'Title', content, QMessageBox.Yes | QMessageBox.Cancel)

    def onButtonClick1(self):
        qApp = QApplication.instance()
        qApp.quit()

    def onButtonClick2(self):
        self.contents.clear()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())
